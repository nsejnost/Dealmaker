"""Read loan parameters and cashflows from the Ginnie Project Loan Maker workbook."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.loan import (
    CashflowRow,
    LoanInput,
    PricingInput,
    PricingType,
)

SHEET_NAME = "CL Model"

# Excel epoch for serial date conversion (1900 system, with the leap-year bug)
_EXCEL_EPOCH = date(1899, 12, 30)


def _serial_to_date(serial: int | float) -> date:
    """Convert an Excel serial number to a Python date."""
    return _EXCEL_EPOCH + timedelta(days=int(serial))


def _to_date(value: Any) -> date:
    """Coerce an Excel cell value to a Python date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return _serial_to_date(value)
    raise ValueError(f"Cannot convert {value!r} to date")


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    return float(value)


def _to_int(value: Any) -> int:
    if value is None:
        return 0
    return int(round(float(value)))


def read_loan_inputs(ws: Worksheet) -> LoanInput:
    """Read loan input parameters from the CL Model sheet.

    Cell mapping (column F, rows 8-19):
      F8  = dated_date
      F9  = first_settle
      F10 = delay
      F11 = original_face
      F13 = coupon_net
      F14 = wac_gross
      F15 = wam
      F16 = amort_wam
      F17 = io_period
      F18 = balloon
      F19 = seasoning
    """
    return LoanInput(
        dated_date=_to_date(ws["F8"].value),
        first_settle=_to_date(ws["F9"].value),
        delay=_to_int(ws["F10"].value),
        original_face=_to_float(ws["F11"].value),
        coupon_net=_to_float(ws["F13"].value),
        wac_gross=_to_float(ws["F14"].value),
        wam=_to_int(ws["F15"].value),
        amort_wam=_to_int(ws["F16"].value),
        io_period=_to_int(ws["F17"].value),
        balloon=_to_int(ws["F18"].value),
        seasoning=_to_int(ws["F19"].value),
    )


def read_pricing_inputs(ws: Worksheet) -> PricingInput:
    """Read pricing parameters from the CL Model sheet.

    Cell mapping:
      I8  = pricing_input (price or yield value)
      I9  = pricing_type  ("Price" or "Yield")
      I10 = settle_date
      I11 = curve_date
    """
    raw_type = str(ws["I9"].value or "Price").strip()
    pricing_type = PricingType.YIELD if raw_type.lower() == "yield" else PricingType.PRICE

    return PricingInput(
        pricing_type=pricing_type,
        pricing_input=_to_float(ws["I8"].value),
        settle_date=_to_date(ws["I10"].value),
        curve_date=_to_date(ws["I11"].value),
    )


def read_cashflows(ws: Worksheet) -> list[CashflowRow]:
    """Read the 121-row cashflow table from rows 32-152, columns E-R.

    Column mapping:
      E = month, F = date_serial, G = cf_date_serial, H = year_frac,
      I = beg_bal, J = pmt_to_agy, K = int_to_inv, L = int_to_agy,
      M = reg_prn, N = rem_prn, O = balloon_pay, P = end_bal,
      Q = net_prn, R = net_flow
    """
    rows: list[CashflowRow] = []
    for row_num in range(32, 153):  # rows 32 through 152 inclusive
        month_val = ws.cell(row=row_num, column=5).value  # col E
        if month_val is None:
            break

        # Date columns may be actual dates or serial numbers
        date_val = ws.cell(row=row_num, column=6).value  # col F
        cf_date_val = ws.cell(row=row_num, column=7).value  # col G

        def _date_to_serial(val: Any) -> int:
            if isinstance(val, (int, float)):
                return int(val)
            if isinstance(val, datetime):
                return (val.date() - _EXCEL_EPOCH).days
            if isinstance(val, date):
                return (val - _EXCEL_EPOCH).days
            return 0

        rows.append(CashflowRow(
            month=_to_int(month_val),
            date_serial=_date_to_serial(date_val),
            cf_date_serial=_date_to_serial(cf_date_val),
            year_frac=_to_float(ws.cell(row=row_num, column=8).value),   # H
            beg_bal=_to_float(ws.cell(row=row_num, column=9).value),     # I
            pmt_to_agy=_to_float(ws.cell(row=row_num, column=10).value), # J
            int_to_inv=_to_float(ws.cell(row=row_num, column=11).value), # K
            int_to_agy=_to_float(ws.cell(row=row_num, column=12).value), # L
            reg_prn=_to_float(ws.cell(row=row_num, column=13).value),    # M
            rem_prn=_to_float(ws.cell(row=row_num, column=14).value),    # N
            balloon_pay=_to_float(ws.cell(row=row_num, column=15).value),# O
            end_bal=_to_float(ws.cell(row=row_num, column=16).value),    # P
            net_prn=_to_float(ws.cell(row=row_num, column=17).value),    # Q
            net_flow=_to_float(ws.cell(row=row_num, column=18).value),   # R
        ))
    return rows


def read_analytics_outputs(ws: Worksheet) -> dict[str, float]:
    """Read computed analytics from the Excel workbook.

    Cell mapping:
      I13 = price, I14 = accrued, I15 = yield_pct,
      I16 = j_spread (bp), I17 = wal,
      I18 = modified_duration, I19 = convexity,
      I20 = risk_dpdy, I21 = tsy_rate_at_wal
    """
    return {
        "price": _to_float(ws["I13"].value),
        "accrued": _to_float(ws["I14"].value),
        "yield_pct": _to_float(ws["I15"].value),
        "j_spread": _to_float(ws["I16"].value),
        "wal": _to_float(ws["I17"].value),
        "modified_duration": _to_float(ws["I18"].value),
        "convexity": _to_float(ws["I19"].value),
        "risk_dpdy": _to_float(ws["I20"].value),
        "tsy_rate_at_wal": _to_float(ws["I21"].value),
    }


def parse_workbook(file_bytes: bytes) -> dict:
    """Parse a Ginnie Project Loan Maker workbook and return extracted data.

    Returns a dict with keys:
      - loan: LoanInput dict
      - pricing: PricingInput dict
      - cashflows: list of CashflowRow dicts
      - analytics: dict of analytics outputs from the workbook
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]
        loan = read_loan_inputs(ws)
        pricing = read_pricing_inputs(ws)
        cashflows = read_cashflows(ws)
        analytics = read_analytics_outputs(ws)
    finally:
        wb.close()

    return {
        "loan": loan.model_dump(mode="json"),
        "pricing": pricing.model_dump(mode="json"),
        "cashflows": [cf.model_dump() for cf in cashflows],
        "analytics": analytics,
    }
